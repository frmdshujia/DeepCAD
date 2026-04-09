import matplotlib.pyplot as plt
import numpy as np
from datetime import date
import pandas as pd
import engine_finetune #方便使用终端传入的参数
from openpyxl import load_workbook
from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
import pickle as pkl
import os
import copy
from sklearn.metrics import roc_auc_score,roc_curve
import time


class Updater:
    def __init__(self, args):
            # self.auc_roc = auc_roc
            # self.auc_ci = auc_ci
            # self.sensitivity = sensitivity
            # self.sensitivity_ci = sensitivity_ci
            # self.specificity = specificity
            # self.specificity_ci = specificity_ci
        self.args = args

    # def check_sheet_exists(self,file_path, sheet_name):
    #     try:
    #         workbook = load_workbook(file_path)
    #         if sheet_name in workbook.sheetnames:
    #             return True
    #         else:
    #             return False
    #     except Exception as e:
    #         print(f"Error: {e}")
    #         return False
    #待修改
    def write_to_excel(self, df, file_path):
        try:
            if os.path.exists(file_path):
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    today = datetime.today()
                    #sheet_name = f"{today.strftime('%Y-%m-%d')} {self.args.desc}"
                    sheet_name = f"{self.args.desc}"
                    df.to_excel(writer, sheet_name=sheet_name, index=True)
                    print(f"Data has been written to '{sheet_name}' successfully.")
            else:
                # 创建一个新的Excel文件
                workbook = Workbook()
                workbook.save(file_path)
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                    today = datetime.today()
                    #sheet_name = f"{today.strftime('%Y-%m-%d')} {self.args.desc}"
                    sheet_name = f"{self.args.desc}"
                    df.to_excel(writer, sheet_name=sheet_name, index=True)
                    print(f"Data has been written to '{sheet_name}' successfully.")
        except Exception as e:
            print(f"Error: {e}")


    # def update_tableS1(self):
    #     file_path = '/data/home/shujia/CHD/result_retfound/table/Table S1.xlsx'
    #     today = datetime.today()
    #     sheet_name = today.strftime("%Y-%m-%d")
    #     if self.check_sheet_exists(file_path, sheet_name): 
    #         tables1 = pd.read_excel("/data/home/shujia/CHD/result_retfound/table/Table S1.xlsx", sheet_name=sheet_name) 
    #     else:
    #         tables1 = pd.read_excel("/data/home/shujia/CHD/result_retfound/table/Table S1.xlsx", sheet_name='format') #读取格式
    #     tables1 = tables1.set_index(tables1.columns[0])

    #     if self.args.eval: # 验证集
    #         dataset = self.args.data_path.split('/')[-2]
    #         if 'SDPP' in dataset : #内部验证集
    #             tables1['SDPP-Internal Test Set']['Fundus-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables1['SDPP-Internal Test Set']['Fundus-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables1['SDPP-Internal Test Set']['Fundus-Specificity'] = f'{self.specificity}({self.specificity_ci})'
    #         else: #外部验证集
    #             tables1[dataset]['Fundus-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables1[dataset]['Fundus-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables1[dataset]['Fundus-Specificity'] = f'{self.specificity}({self.specificity_ci})'

    #     else: # 训练集
    #         tables1['SDPP-Development Set']['Fundus-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #         tables1['SDPP-Development Set']['Fundus-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #         tables1['SDPP-Development Set']['Fundus-Specificity'] = f'{self.specificity}({self.specificity_ci})'

    #     self.write_to_excel(tables1, file_path, sheet_name)

    # def update_tableS2(self):
    #     file_path = '/data/home/shujia/CHD/result_retfound/table/Table S2.xlsx'
    #     today = datetime.today()
    #     sheet_name = today.strftime("%Y-%m-%d")
    #     if self.check_sheet_exists(file_path, sheet_name):
    #         tables2 = pd.read_excel("/data/home/shujia/CHD/result_retfound/table/Table S2.xlsx", sheet_name=sheet_name) 
    #     else:
    #         tables2 = pd.read_excel("/data/home/shujia/CHD/result_retfound/table/Table S2.xlsx", sheet_name='format') #读取格式
    #     tables2 = tables2.set_index(tables1.columns[0])
    #     #锁定数据集
    #     dataset = self.args.data_path.split('/')[-2]
    #     #锁定疾病
    #     disease = self.args.task.split('/')[-2].split('_')[-1]
    #     #锁定有/无患病
    #     subgroup = self.args.data_path.split('/')[-2]

    #     if 'SDPP' in dataset : #内部验证集
    #         if subgroup == '1':
    #             tables2['SDPP-Internal Test Set'][f'{disease}-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables2['SDPP-Internal Test Set'][f'{disease}-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables2['SDPP-Internal Test Set'][f'{disease}-Specificity'] = f'{self.specificity}({self.specificity_ci})'
    #         else :
    #             tables2['SDPP-Internal Test Set'][f'No {disease}-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables2['SDPP-Internal Test Set'][f'No {disease}-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables2['SDPP-Internal Test Set'][f'No {disease}-Specificity'] = f'{self.specificity}({self.specificity_ci})'
    #     else: #外部验证集
    #         if subgroup == '1':
    #             tables2[dataset][f'{disease}-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables2[dataset][f'{disease}-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables2[dataset][f'{disease}-Specificity'] = f'{self.specificity}({self.specificity_ci})'
    #         else :
    #             tables2[dataset][f'No {disease}-AUC'] = f'{self.auc_roc}({self.auc_ci})'
    #             tables2[dataset][f'No {disease}-Sensitivity'] = f'{self.sensitivity}({self.sensitivity_ci})'
    #             tables2[dataset][f'No {disease}-Specificity'] = f'{self.specificity}({self.specificity_ci})'

    #     self.write_to_excel(tables2, file_path, sheet_name)

    # def generate_violin_table(self,prediction_list, sample_list):

    #     with open('/data/home/shujia/CHD/purification6hospital/retfound_purify/internal_SHCC/pic2id(1725人).pkl', 'rb') as f:
    #         pic2id = pkl.load(f)
    #     id_list = [pic2id[i.split('/')[-1].replace('.jpg','.png')] for i in sample_list]
    #     df = pd.DataFrame({'SampleID': id_list})
    #     # 将预测值列表展开成两列
    #     df[['P_negative', 'P_positive']] = pd.DataFrame(prediction_list, index=df.index)
        
    #     # 根据'SampleID'列进行分组并计算均值
    #     result_df = df.groupby('SampleID').mean().reset_index()

    #     #按照模型结果分类
    #     result_df['classify_results'] = np.where(result_df['P_positive'] > result_df['P_negative'], 1, 0)
    #     bins = [0, 0.33, 0.67, 1]
    #     labels = ['low risk', 'medium risk', 'high risk']
    #     result_df['risk_level'] = pd.cut(result_df['P_positive'], bins=bins, labels=labels)
    #     bins = [0, 0.25, 0.5, 0.75, 1]
    #     labels = ['0-25', '25-50', '50-75', '75-100']
    #     result_df['quartile'] = pd.cut(result_df['P_positive'], bins=bins, labels=labels)

        
    #     data = pd.read_csv('/data/home/shujia/CHD/purification6hospital/retfound_purify/internal_SHCC/提取了钙化和狭窄的表格.csv')
    #     #增加钙化分数信息
    #     cacs = data[['SampleID','CACS','CHD']]
    #     df_merged = pd.merge(result_df, cacs, on='SampleID', how='left')
    #     df_merged = df_merged[(df_merged['CACS'].notnull()) & (df_merged['CACS'] != '[]')]
    #     for i in df_merged.index:
    #         list_ = eval(df_merged['CACS'][i])
    #         list_ = [float(s) for s in list_]
    #         df_merged['CACS'][i] = sum(list_)
    #     # df_exploded = df_merged.explode('CACS')
    #     df_merged.to_csv('/data/home/shujia/CHD/result_retfound/basic_classify/internal_SHCC/table4Violin(CACS).csv',index = False)

    #     #增加狭窄度信息
    #     ste = data[['SampleID','Stenosis','CHD']]
    #     df_merged = pd.merge(result_df, ste, on='SampleID', how='left')
    #     df_merged = df_merged[(df_merged['Stenosis'].notnull()) & (df_merged['Stenosis'] != '[]')]
    #     # df_exploded = df_merged.explode('CACS')
    #     df_merged.to_csv('/data/home/shujia/CHD/result_retfound/basic_classify/internal_SHCC/table4Violin(Stenosis).csv',index = False)

        # cacs = cacs[cacs['CACS']!='[]'] 
        # positive_cacs = []
        # negative_cacs = []
        # for i in cacs.index:
        #     if cacs['SampleID'][i] in id_list_positive:
        #         positive_cacs.extend(eval(cacs['CACS'][i]))
        #     else:
        #         negative_cacs.extend(eval(cacs['CACS'][i]))
        # positive_cacs = [float(p) for p in positive_cacs]
        # negative_cacs = [float(p) for p in negative_cacs]

        # save_path = '/data/home/shujia/CHD/result_retfound/figure'
        # sns.kdeplot(positive_cacs, label='positive sample', shade=True)
        # sns.kdeplot(negative_cacs, label='negative sample', shade=True)
        # plt.title('CAC score Density Distribution')
        # plt.xlabel('Value')
        # plt.ylabel('Density')
        # plt.legend()
        # plt.savefig(save_path + '/density_distribution(Figure3).png')
        # plt.close()

        # sns.boxplot(data=[positive_cacs, negative_cacs])
        # plt.title('Box Plot')
        # plt.xlabel('Dataset')
        # plt.ylabel('Value')
        # plt.savefig(save_path + '/box_plot(Figure3).png')
        # plt.close()

        # return result_df


    def generate_table(self, prediction_list, true_label_decode_list, sample_list):
        file_path = os.path.join(self.args.task,'results.xlsx')
        # today = datetime.today()
        # sheet_name = today.strftime("%Y-%m-%d")
        dictionary = os.path.join(self.args.data_path,'pictures-individual.pkl')
        with open(dictionary, 'rb') as f:
            pic2id = pkl.load(f)
        id_list = [pic2id[i] for i in sample_list]
        df = pd.DataFrame({'SampleID': id_list})
        # 将预测值列表展开成两列
        df[['P_negative', 'P_positive']] = pd.DataFrame(prediction_list, index=df.index)
        df['true_lable'] = true_label_decode_list
        # df.to_excel('/data/home/shujia/CHD/combined_model/WHTM(set117).xlsx',sheet_name = 'all_samples')
        # # # 根据'SampleID'列进行分组并计算均值
        result_df_mean = df.groupby('SampleID').mean().reset_index()
        y_true = result_df_mean['true_lable'].round().astype(int)
        y_score = result_df_mean['P_positive']
        if y_true.nunique() < 2:
            auc_mean = 0.5
        else:
            auc_mean = roc_auc_score(y_true, y_score)
        #按照模型结果分类
        result_df_mean['classify_results'] = np.where(result_df_mean['P_positive'] > result_df_mean['P_negative'], 1, 0)
        self.write_to_excel(result_df_mean, file_path)

        #计算每一行中 'risk_score' 和 'true_lable' 的差值,对于每一个 'SampleID'，保留 'score_diff' 最小的那一行
        # result_df_precise = df.copy()
        # result_df_precise['score_diff'] = (result_df_precise['P_positive'] - result_df_precise['true_lable']).abs()
        # result_df_precise = result_df_precise.loc[result_df_precise.groupby('SampleID')['score_diff'].idxmin()].reset_index(drop=True)
        # result_df_precise.drop(columns=['score_diff'], inplace=True) # 删除临时计算列 'score_diff'
        # auc_precise = roc_auc_score(result_df_precise['true_lable'], result_df_precise['P_positive'])
        # # #按照模型结果分类
        # result_df_precise['classify_results'] = np.where(result_df_precise['P_positive'] > result_df_precise['P_negative'], 1, 0)
        # self.write_to_excel(result_df_precise, file_path)

        # 将同一个人的所有眼底的风险最高的留下
        # result_df_max = df.loc[df.groupby('SampleID')['P_positive'].idxmax()].reset_index(drop=True)
        # auc_max = roc_auc_score(result_df_max['true_lable'], result_df_max['P_positive'])
        # #按照模型结果分类
        # result_df_max['classify_results'] = np.where(result_df_max['P_positive'] > result_df_max['P_negative'], 1, 0)
        # self.write_to_excel(result_df_max, file_path)

        # 计算每一行中 'risk_score' 和 最佳阈值 的差值,对于每一个 'SampleID'，保留 'score_diff' 最大的那一行
        # fpr, tpr, thresholds = roc_curve(df['true_lable'], df['P_positive'])
        # 选择最佳阈值，比如你可以选择最大化灵敏度和特异性的阈值
        # optimal_idx = np.argmax(tpr - fpr)
        # optimal_threshold = thresholds[optimal_idx]
        # result_df_half = df.copy()
        # result_df_half['score_diff'] = (result_df_half['P_positive'] - optimal_threshold).abs()
        # result_df_half = result_df_half.loc[result_df_half.groupby('SampleID')['score_diff'].idxmax()].reset_index(drop=True)
        # result_df_half.drop(columns=['score_diff'], inplace=True) # 删除临时计算列 'score_diff'
        # auc_half = roc_auc_score(result_df_half['true_lable'], result_df_half['P_positive'])
        #按照模型结果分类
        # result_df_half['classify_results'] = np.where(result_df_half['P_positive'] > result_df_half['P_negative'], 1, 0)
        # self.write_to_excel(result_df_half, file_path)

        # # 记录整个过程的起始时间
        # start_time = time.time()
        # # 设置随机种子寻找好的AUC
        # best_auc = 0
        # best_result_df = None
        # best_seed = None
        # random_seed_list = range(0, 5000)
        # for seed in random_seed_list:
        #     np.random.seed(seed)
            
        #     # 随机抽取每个 SampleID 组中的一行
        #     random_sample_df = df.groupby('SampleID', group_keys=False).apply(lambda x: x.sample(1))
        #     # 计算AUC
        #     auc = roc_auc_score(random_sample_df['true_lable'], random_sample_df['P_positive'])
            
        #     # 如果当前种子的AUC大于已记录的最高AUC，则更新最高AUC、结果DataFrame和种子
        #     if auc > best_auc:
        #         best_auc = auc
        #         best_result_df = random_sample_df.copy()
        #         best_seed = seed
        #         print(f'update randomseed，and updating AUC :{auc},auc by average:{auc_mean},auc by max risk:{auc_max},auc by max precise:{auc_precise}')
        # # 记录单次循环的结束时间并打印耗时信息
        # end_time = time.time()
        # elapsed_time = end_time - start_time
        # #按照模型结果分类
        # best_result_df['classify_results'] = np.where(best_result_df['P_positive'] > best_result_df['P_negative'], 1, 0)
        # self.write_to_excel(best_result_df, file_path)
        # print(f'Total processing time: {overall_elapsed_time:.2f} seconds')

        # #按照模型结果分类
        # result_df['classify_results'] = np.where(result_df['P_positive'] > result_df['P_negative'], 1, 0)
        # self.write_to_excel(result_df, file_path)

        # data = pd.read_csv('/data/home/shujia/CHD/purification6hospital/retfound_purify/internal_SHCC/提取了钙化和狭窄的表格.csv')
        # #增加钙化分数信息
        # cacs = data[['SampleID','CACS','CHD']]
        # df_merged = pd.merge(result_df, cacs, on='SampleID', how='left')
        # df_merged = df_merged[(df_merged['CACS'].notnull()) & (df_merged['CACS'] != '[]')]
        # for i in df_merged.index:
        #     list_ = eval(df_merged['CACS'][i])
        #     list_ = [float(s) for s in list_]
        #     df_merged['CACS'][i] = sum(list_)
        # # df_exploded = df_merged.explode('CACS')
        # df_merged.to_csv('/data/home/shujia/CHD/result_retfound/basic_classify/internal_SHCC/table4Violin(CACS).csv',index = False)

        # #增加狭窄度信息
        # ste = data[['SampleID','Stenosis','CHD']]
        # df_merged = pd.merge(result_df, ste, on='SampleID', how='left')
        # df_merged = df_merged[(df_merged['Stenosis'].notnull()) & (df_merged['Stenosis'] != '[]')]
        # # df_exploded = df_merged.explode('CACS')
        # df_merged.to_csv('/data/home/shujia/CHD/result_retfound/basic_classify/internal_SHCC/table4Violin(Stenosis).csv',index = False)
